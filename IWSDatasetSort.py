from tkinter.constants import DISABLED
from typing import Text
import PySimpleGUI as sg
from openpyxl import Workbook
import os
import shutil
import numpy as np
import re

from openpyxl.reader.excel import load_workbook

sg.theme('DarkAmber')
def GUI(*args):
    frame_layout = [ [sg.Listbox(values=['胶质瘤85例已标注','脑膜瘤103例已标注','脑膜瘤110例未标注'],size=(30,7))]]
    layout = [[sg.Text('IWSDataset')], 
              [sg.Text('根目录',size=(15,1),key='-Folder-'),sg.InputText(key='-SourcePath-'),sg.FolderBrowse('FolderBrowse', size=(10,1))],
              [sg.Text('选择文件',size=(15,1),key='-file-'),sg.InputText(),sg.FilesBrowse('FileBrowse', size=(10,1), key=('-FilesBrowse-'),initial_folder='F:/PythonProject/IWSDatasetSort')],
              [sg.Text('所有病例',size=(15,1)),sg.Combo(['所有','已标注','未标注'],default_value='所有', size=(10,1),key=('-UntaggedDiseazeName-'))],
              [sg.Text('病种类型',size=(15,1)),sg.Combo(['脑膜瘤','胶质','CPA表皮样囊肿','ACTH腺瘤'],default_value = '脑膜瘤', size=(10,1),key='-TaggedDiseazeName-'),
              sg.Text('因存在“胶质细胞瘤”的命名,所以检索时以“胶质”作为检索项')],
              [sg.Text('输出信息',size=(15,1)),sg.Combo(['数量','所在行','文件夹名称'],default_value='数量',size=(10,1),key='-LogData-')],
              [sg.Frame('已提交数据',frame_layout,font='Any 12',title_color='blue')],
              [sg.Checkbox('拷贝检索数据至输出文件夹',default=False,disabled=False,key=('-CopyCheckbox-'),change_submits=True)],
              [sg.Text('输出文件夹',size=(15,1)),sg.Input(key=('-TargetPath-')),sg.FolderBrowse('OutputFolder',size=(10,1))],
              [sg.Button(button_text = 'Search',key='-Submit-'),sg.Button('Copy',key=('-Copy-'),disabled=True,visible=False)],
              [sg.Cancel()]
             ]

    window = sg.Window('IWS DATASET',layout,location=(0,0))

    while True:
        event,values=window.read()
        print(event)
        if event in (sg.WIN_CLOSED,'Cancel'):
            break

        if values['-CopyCheckbox-']:
            window['-Submit-']('Copy')
        else:
            window['-Submit-']('Search')
        
        if event == '-Submit-':
            if values['-CopyCheckbox-']:
                ReadXlsxFile(path=values['-FilesBrowse-'],UntaggedDiseazeName=values['-UntaggedDiseazeName-'], TaggedDiseazeName=values['-TaggedDiseazeName-'], logData=values['-LogData-'],
                enableCopy=values['-CopyCheckbox-'],sourcePath=values['-SourcePath-'],targetPath=values['-TargetPath-'])
            else:
                ReadXlsxFile(path=values['-FilesBrowse-'],UntaggedDiseazeName=values['-UntaggedDiseazeName-'], TaggedDiseazeName=values['-TaggedDiseazeName-'], logData=values['-LogData-'])

    window.close()

# 检索表格中影像数据
def ReadXlsxFile(path=None,UntaggedDiseazeName=None,TaggedDiseazeName=None,logData=None,enableCopy=False,sourcePath=None,targetPath=None):
    wb = Workbook()
    if os.path.exists(path):
        wb = load_workbook(path)
        sheet = wb.active
    else:
        sg.Print("请确保已选择文件")
    
    # 根据表格数据检索病种类型
    if UntaggedDiseazeName == '所有':
        diseazeNameList = sheet['E']
    elif UntaggedDiseazeName == '已标注':
        diseazeNameList = sheet['G']
    elif UntaggedDiseazeName == '未标注':
        return
    

    # 根据选择的病种名称进行检索
    num = 0
    for i in diseazeNameList:
        if str(i.value).find(TaggedDiseazeName) >= 0 :
            num+=1
            if logData == '数量':
                sg.Print(num)
            if logData =='所在行':
                sg.Print(i.row)
            if logData == '文件夹名称':
                sg.Print(sheet['A'+str(i.row)].value)
            if enableCopy:
                folderName = '/' + sheet['A'+str(i.row)].value
                shutil.copytree(sourcePath +folderName,targetPath+folderName)
GUI()