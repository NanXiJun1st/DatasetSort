import os
from tkinter.constants import DISABLED
import PySimpleGUI as sg
from PySimpleGUI.PySimpleGUI import WINDOW_CLOSED
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

sg.theme('DarkAmber')

def GUI(*args):
    layout = [
        [sg.Text('新增数据文件夹:',size=(15,1)),sg.InputText(key='-FolderPath-',enable_events=True),sg.FolderBrowse(initial_folder='C:/Users/VertecxGd/Desktop/')],
        [sg.Text('数据统计表:',size=(15,1)),sg.InputText(key='-FilePath-'),sg.FilesBrowse(file_types = (('Excel Files','*.xlsx'),),initial_folder='C:/Users/VertecxGd/Desktop/')],
        [sg.Text('开始列',size=(15,1)),sg.Spin([i for i in range(1,10480)],initial_value = 1, size=(5,1), disabled=True, key='-StartRow-',enable_events=True),
        sg.Checkbox('确认',key='-Checkbox-',enable_events=True,default=False)],
        [sg.Submit(),sg.Cancel()]
        ]

    window = sg.Window('CopyFolderNameToExcel',layout,location=(0,0))

    while True:
        event, values = window.read()
        if event in (WINDOW_CLOSED,'Cancel'):
            break
        if event == '-FolderPath-':
            for dir in os.listdir(values['-FolderPath-']):
                sg.Print(dir)
        if values['-Checkbox-']:
            window['-StartRow-'].update(disabled=False)
        if event == 'Submit':
            CopyFolderNameAndWriteIn(values['-FolderPath-'],values['-FilePath-'],values['-StartRow-'])
    window.close()

def CopyFolderNameAndWriteIn(folderPath=None,filePath=None,startRow=None):
    """
    :folderPath: 需添加的文件路径
    :filePath：表格文件路径
    :startRow：开始插入行
    """
    wb=Workbook()
    num=startRow

    if len(filePath)>0 & os.path.exists(filePath):
        wb = load_workbook(filePath)
        sheet = wb.active

        # 获取选择文件路径下的第一层文件夹
        dirsList = os.listdir(folderPath)

        # 将第一层的文件夹名称写入Excel
        for dir in dirsList:
            num+=1
            sheet['A'+ str(num)].value = dir
            sg.Print(dir)
        wb.save(filePath)

        # 打开编辑后的Excel
        os.startfile(filePath)
    else:
        sg.Print("请确保已选择文件")

GUI()